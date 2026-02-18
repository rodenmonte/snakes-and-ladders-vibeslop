# Any gratuitous commenting is an artifact of vibe-slopping, best ignored
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from math import log, ceil
from scipy.stats import nbinom

# Helper: median kills for r successes at drop rate p per kill
def median_kills_nbinom(r, p):
    """Median kills to get r drops at probability p per kill."""
    return int(ceil(nbinom.ppf(0.5, r, p)))

# Helper: for "any of N items each at rate p" -> combined rate
def combined_rate(rates):
    """Given a list of per-kill probabilities, return P(at least one)."""
    prob_none = 1
    for r in rates:
        prob_none *= (1 - r)
    return 1 - prob_none

# ============================================================
# DROP RATE & KILL RATE DATABASE
# ============================================================
# Kill rates are in kills/hour (ironman EHB-style estimates for geared players)
# Drop rates are per-kill probability of the specific item

# NOTE: These are best estimates from known OSRS data.
# Items marked with confidence="low" need wiki verification.

tiles = []

def add_tile(tile_num, description, median_hours, notes="", confidence="high", category="obtain"):
    tiles.append({
        "tile": tile_num,
        "description": description,
        "median_hours": round(median_hours, 2),
        "notes": notes,
        "confidence": confidence,
        "category": category,
    })

def add_movement(tile_num, description, target):
    tiles.append({
        "tile": tile_num,
        "description": description,
        "median_hours": 0,
        "notes": f"Move to tile #{target}",
        "confidence": "n/a",
        "category": "movement",
    })

def add_free(tile_num, description):
    tiles.append({
        "tile": tile_num,
        "description": description,
        "median_hours": 0,
        "notes": "Free tile",
        "confidence": "n/a",
        "category": "free",
    })

# ============================================================
# REUSABLE ESTIMATES FOR RECURRING TILE TYPES
# ============================================================

class BossOrRaidForUnique:
    def __init__(self, name, unique_rate, ehb):
        self.name = name
        self.unique_rate = unique_rate
        self.ehb = ehb
        self.median_kc = ceil(log(0.5) / log(1 - self.unique_rate))
        self.hours_to_unique = self.median_kc / self.ehb

    def hours_for_two_uniques(self):
        return median_kills_nbinom(2, self.unique_rate) / self.ehb

# --- Slayer Boss Drop (best option) ---
# 1/256 unqiue at shellbane gryphon, 95 ehb, 178 median on rate, 1.8 hour/unique
# 62/3000 unique at GG's, 34 kph, 34 median to go on rate (really!), 1 hour/unique
# 1/100 unique at sire, 39 EHB, 69 median on rate, 2 hour/unique
# 10/3000 unique at kraken (jar/pet/trident), 60 kph, ~210 median, 3.5 hours/unique
# 4/520 unique at cerb, 50 EHB, ~90 kills median, 1.8 hours/unique
# ~10/2000 unique chance at thermy, 80 EHB, ~139 kills median, 1.8 hours/unique
# 10/600 unique at araxxor, 38 kph, 42 median on rate, 1.1 hour/unique
# ~24/2160 unique chance at hydra, 25 EHB, ~62 kills median for unique, 2.5 hours/unique
SLAYER_BOSSES = [
    BossOrRaidForUnique("Shellbane Gryphon", 1/256, 95),
    BossOrRaidForUnique("Grotesque Guardians", 62/3000, 34),
    BossOrRaidForUnique("Abyssal Sire", 1/100, 39),
    BossOrRaidForUnique("Kraken", 10/3000, 60),
    BossOrRaidForUnique("Cerberus", 4/520, 39),
    BossOrRaidForUnique("Thermonuclear Smoke Devil", 10/2000, 80),
    BossOrRaidForUnique("Araxxor", 10/600, 38),
    BossOrRaidForUnique("Alchemical Hydra", 24/2160, 25)
]

SLAYER_BOSS_1X_HOURS = min(map(lambda boss: boss.hours_to_unique, SLAYER_BOSSES))
SLAYER_BOSS_2X_HOURS = min(map(lambda boss: boss.hours_for_two_uniques(), SLAYER_BOSSES))

# --- Raid Drop (any raid, unique table) ---
# While team purple rates go up in group raids, individual doesn't (ignore ToB)
# Solo ToA: 30 minute 300's need 16 raids median (4.3% chance of purple), 8 hours
# Solo ToA: 40 minute 500's 7.6% purple chance 9 kc median (7.6% purple chance), 6 hours
# Solo CM: 93k points per hour, 10.72% purple per hour, (10.72% purp chance), 6 hours
# Trio ToB: 1/9.1 unique chance, 1/27.3 (3.6%) solo, 19 kills needed median, 20 minutes/raid = 6.3 hours
# HMT Trio: 1/7.7, 1/23.1 solo, 16 kills needed median, 24 minutes/raid = 6.4 hours
RAIDS = [
    BossOrRaidForUnique("300 ToA (30 minute solo)", 0.043, 2),
    BossOrRaidForUnique("500 ToA (40 minute solo)", 0.076, 1.5),
    BossOrRaidForUnique("CoX (93k points/hour)", 0.1072, 1),
    BossOrRaidForUnique("Trio ToB", 1/27.3, 3),
    BossOrRaidForUnique("Trio HMT", 1/23.1, (60 / 24))
]

RAID_1X_HOURS = min(map(lambda raid: raid.hours_to_unique, RAIDS))
RAID_2X_HOURS = min(map(lambda raid: raid.hours_for_two_uniques(), RAIDS))

# --- Barrows Unique ---
# 24 items, 1/17.42 for any unique per chest (with max reward potential)
# ~15 chests/hr with Barrows tele + max gear
# Median for 3 uniques: nbinom(3, 1/17.42)
p_barrows = 1/17.42
barrows_kph = 15
med_3_barrows = int(ceil(nbinom.ppf(0.5, 3, p_barrows)))
med_4_barrows = int(ceil(nbinom.ppf(0.5, 4, p_barrows)))
med_5_barrows = int(ceil(nbinom.ppf(0.5, 5, p_barrows)))
BARROWS_3X_HOURS = round(med_3_barrows / barrows_kph, 2)
BARROWS_4X_HOURS = round(med_4_barrows / barrows_kph, 2)
BARROWS_5X_HOURS = round(med_5_barrows / barrows_kph, 2)

# --- DK Rings (Warrior, Berserker, Seer, Archer) ---
# I assume we can get e.g. 2 or 3 b rings and that will count...?
p_dk_ring = 0.04639
dk_kph = 66 # 22 kills of each one/hour https://oldschool.runescape.wiki/w/Money_making_guide/Killing_Dagannoth_Kings_(Solo_tribrid)
med_3_dk = int(ceil(nbinom.ppf(0.5, 3, p_dk_ring)))
DK_3X_HOURS = round(med_3_dk / dk_kph, 2)

# --- Moons of Peril ---
# 18 kph https://oldschool.runescape.wiki/w/Money_making_guide/Moons_of_Peril
# 1/19 for unique https://oldschool.runescape.wiki/w/Lunar_Chest
p_moons = 1/19
moons_kph = 18
def moons_hours(n):
    return round(int(ceil(nbinom.ppf(0.5, n, p_moons))) / moons_kph, 2)

# --- Hueycoatl ---
# 1/70 for a unique in a trio
p_huey = 1/70
huey_kph = 20
def huey_hours(n):
    return round(int(ceil(nbinom.ppf(0.5, n, p_huey))) / huey_kph, 2)

# --- Doom of Mokhaiotl ---
# Claiming wave 8
p_doom = 1/50 # Odds of a unique by wave 8
doom_kph = 6
def doom_hours(n):
    return round(int(ceil(nbinom.ppf(0.5, n, p_doom))) / doom_kph, 2)

# --- GWD Drop (unique, no shards) ---
# Fastest: probably Kree'arra or Zilyana
# Each GWD boss has ~1/127 for each unique (hilt, or specific armor piece)
# Kree has 4 uniques each ~1/127 -> combined ~1/32. ~25 kills/hr
# Actually each GWD boss drops from a shared table:
# Armadyl: helmet 1/381, chestplate 1/381, chainskirt 1/381, hilt 1/508 -> combined ~1/109
# Wait those rates are wrong. Let me reconsider.
# GWD unique rates are approximately 1/128 per item for armor, 1/512 for hilt
# Kree'arra: Armadyl helmet/chest/skirt each 1/384, Armadyl hilt 1/512
# Combined: 3/384 + 1/512 = ~1/103
# ~30 kills/hr with good team/gear
# Median: ceil(log(0.5)/log(102/103)) = ~71 kills / 30 = 2.4 hr
p_gwd = 1/103
gwd_kph = 25
med_1_gwd = ceil(log(0.5) / log(1 - p_gwd))
GWD_1X_HOURS = round(med_1_gwd / gwd_kph, 2)

# --- Crystal Armour Seed ---
# From Gauntlet/Corrupted Gauntlet. CG: 1/50, ~6 completions/hr
# Regular Gauntlet: 1/2000(?) No, crystal armour seed is 1/50 from CG
# Median: ceil(log(0.5)/log(49/50)) = 34 / 6 = 5.7 hr
# Can also get from Zalcano (1/200 from Zalcano, ~30 kph) -> median 138/30 = 4.6 hr
# And from CG 1/50 at 6/hr = 5.7 hr
# And regular Gauntlet: 1/2000 - way too slow
# Zalcano might actually be slightly faster but let's say ~5 hours
p_cg_crystal = 1/50
cg_kph = 6
med_1_crystal = ceil(log(0.5) / log(1 - p_cg_crystal))
CRYSTAL_SEED_1X_HOURS = round(med_1_crystal / cg_kph, 2)

# --- Ecumenical Key ---
# Just kill goblins and imps free tile
ECUMENICAL_3X_HOURS = 1

# --- DT2 Boss Drop ---
# Duke Sucellus, Whisperer, Vardorvis, Leviathan
# Each has uniques + secondary table. Best combined rate:
# Vardorvis: ~40 kph, virtus pieces 1/1500, executioner's axe head 1/3000, 
#   chromium ingot 1/150 (secondary) -> with secondary, ~1/136 combined
# Duke: ~25 kph, various uniques
# The task says "unique AND secondary unique tables" so secondaries count too
# Vardorvis with secondary: chromium ingot 1/150 alone makes this fast
# 40 kph, ~1/150 for ingot -> median ~104 kills / 40 = 2.6 hr for just ingot
DT2_1X_HOURS = 2.5

# --- Zulrah Unique (include mutagens) ---
# Zulrah uniques: Tanzanite fang, Magic fang, Serpentine visage each 1/512
# Mutagens: each 1/6553 (magma, tanzanite)
# Combined unique (including mutagens): 3/512 + 2/6553 ≈ 1/167
# ~35 kills/hr
# Median: ceil(log(0.5)/log(166/167)) = 116 / 35 = 3.3 hr
# Actually Zulrah uniques (from unique table) are 1/512 each for the 3 items
# and 1/3277 each for the 2 mutagens (not 1/6553, that might be outdated)
# Let me use: combined = 3/512 + 2/3277 ≈ 0.00586 + 0.00061 = 0.00647 ≈ 1/155
p_zulrah = 1/155
zulrah_kph = 35
med_1_zulrah = ceil(log(0.5) / log(1 - p_zulrah))
ZULRAH_1X_HOURS = round(med_1_zulrah / zulrah_kph, 2)

# --- Vorkath unique ---
# Vorkath Head 1/50, Jar of Decay 1/2000, Dragonbone Necklace 1/1000, 
# Draconic Visage 1/5000, Vorki 1/3000, Skeletal Visage 1/5000
# Combined: 1/50 + 1/1000 + 1/2000 + 1/5000 + 1/3000 + 1/5000 ≈ 0.02 + 0.001 + 0.0005 + 0.0002 + 0.000333 + 0.0002 = ~0.0222 ≈ 1/45
# ~30 kills/hr
# Median: 31/30 ≈ 1.0 hr
# But wait, Vorkath's head is guaranteed at 50kc. Most of the probability mass is the head.
# Since it says "from Vorkath" specifically, the head at 1/50 makes this very fast.
VORKATH_UNIQUE_HOURS = 2.0

# --- Tempoross (100 Soaked Pages or 1x unique) ---
# Soaked pages are common - roughly 5-8 per completion
# ~12 completions/hr, so ~60-96 pages/hr. 100 pages ≈ 1.3 hours
# Uniques (Tackle Box 1/400, Tome of Water 1/400, Dragon Harpoon 1/800) per completion
# Combined unique: 2/400 + 1/800 = 1/160
# Median 160 kills at 12/hr = ~9.2 hr for unique
# So soaked pages is way faster: ~1.3 hours
TEMPOROSS_HOURS = 1.5  # 100 soaked pages, conservative

# --- Wintertodt (100 Burnt Pages or Tome of Fire or Dragon Axe) ---
# Burnt pages: ~5-8 per crate average, ~12 crates/hr
# 100 pages / 6.5 avg = ~15.4 crates / 12 per hr = ~1.3 hr
# Tome of Fire: 1/1000 per crate, Dragon Axe: 1/10000 per crate
# So burnt pages is by far the fastest option
WINTERTODT_HOURS = 1.5

# --- Abyssal Whip or Unsired ---
# Whip: 1/512 from Abyssal Demons, ~280 kph (barraging)
# 3x whips: nbinom(3, 1/512) median ≈ 1508 kills / 280 = 5.4 hr
# Unsired: 1/100 from Sire, ~28 kph
# 1x unsired: median ~69 kills / 28 = 2.5 hr
# So Unsired is faster for "3x whip OR 1x unsired"
# For "4x whip or 1x unsired" still unsired is faster
WHIP_OR_UNSIRED_3X_HOURS = 2.5  # 1x unsired from Sire
WHIP_OR_UNSIRED_4X_HOURS = 2.5  # still 1x unsired

# --- Champion Scroll ---
# 1/5000 from any champion creature. Can burst goblins/hobgoblins ~500+/hr
# Median: ceil(log(0.5)/log(4999/5000)) = 3466 / 500 = 6.9 hr
CHAMPION_SCROLL_HOURS = 6.9

# --- Marks of Grace (25) ---
# ~15-20 marks/hr on Seers/Ardougne rooftop
# 25 marks / 17.5 per hr = ~1.4 hr
MARKS_25_HOURS = 1.4

ELDER_CUSTODIANS_PER_HOUR = 150

# ============================================================
# NOW BUILD EVERY TILE
# ============================================================

# Tile 1: 5x Scurrius' Spine (1/33, ~40 kph)
p = 1/33; kph = 40
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(1, "Obtain 5x Scurrius' Spine", med/kph, f"1/33 drop, {kph} kph, median {med} kc")

# Tile 2: Tempoross
add_tile(2, "Obtain 100x Soaked Page or 1x Tempoross unique", TEMPOROSS_HOURS, "100 soaked pages fastest (~6-8/permit, 12 permits/hr)")

# Tile 3: 3x DK Ring
add_tile(3, "Obtain 3x DK Ring", DK_3X_HOURS, f"Combined ring rate ~1/21.5 per trio, 15 trios/hr, median {med_3_dk} trios")

# Tile 4: Movement
add_movement(4, "Advance to Tile #11", 11)

# Tile 5: 5x Fresh Crab Claw (crawblaw isle crabs)
p = 1/8; kph = 150
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(5, "Obtain 5x Fresh Crab Claw", med/kph, f"1/8 from level 23 crabclaw isle crabs, {kph} kph, median {med} kc", confidence="high")

# Tile 6: 3x Barronite piece (Barronite Handle/Guard/Head from Camdozaal, ~1/100 each, combined ~1/33)
p = 1/150; kph = 180  # mining golems in camdozaal, rough estimate
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(6, "Obtain 3x Barronite piece", med/kph, f"Barrornite guard 1/150 from chaos golems, ~{kph} kph", confidence="high")

# Tile 7: 3x Mudskipper Hat (from Mogres, 1/32)
p = 1/30; kph = 120
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(7, "Obtain 3x Mudskipper Hat", med/kph, f"1/30 from Mogres, ~{kph} kph, median {med} kc", confidence="high")

# Tile 8: 4x Left Skull Half (from SS Ankous)
p = 1/33; kph = 180
med = int(ceil(nbinom.ppf(0.5, 4, p)))
add_tile(8, "Obtain 4x Left Skull Half", 3.0, f"1/33.33 from S.S. ankous, ~{kph} kph, median {med} kc", confidence="high")

# Tile 9: 5x Broken Antler (Custodian stalkers)
p = 1/20; kph = 150
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(9, "Obtain 5x Broken Antler", med/kph, f"From Custodian Stalkers, ~1/20, ~{kph} kph", confidence="high")

# Tile 10: 3x Mossy Key (from Bryophyta
p = 1/16; kph = 80  # Using burning claws on bryophyta, similar to Obor https://oldschool.runescape.wiki/w/Giant_key, "players can kill Obor 120+ times per hour when using burning claws, giving approximately 8 keys per hour."
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(10, "Obtain 3x Mossy Key", med/kph, f"1/16 from bryophyta off-task, ~{kph} burning claw speccing, median {med} kc")

# Tile 11: 25x Mark of Grace
add_tile(11, "Obtain 25x Mark of Grace", MARKS_25_HOURS, "~17.5 marks/hr on Ardougne")

# Tile 12: 3x Antler Guard (Custodian Stalker)
p = 1/650; kph = ELDER_CUSTODIANS_PER_HOUR
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(12, "Obtain 3x Antler Guard", med/kph, f"Cannoning Elder custodian stalkers, ~{kph} per hour, median {med} kc", confidence="high")

# Tile 13: 1x Squid Beak (from sailing squid)
# https://oldschool.runescape.wiki/w/Squid_beak
# https://oldschool.runescape.wiki/w/Raw_jumbo_squid
p = 1/612; kph = 300 # 300 per hour average btwn comments here https://old.reddit.com/r/2007scape/comments/1qc5p9h/why_do_jumbo_squid_which_heal_17_and_take_69/nzfpnrk/
med = int(ceil(nbinom.ppf(0.5, 1, p)))
add_tile(13, "Obtain 1x Squid Beak", med/kph, f"Catching jumbo squid, ~{kph} per hour, median {med} kc", confidence="high")

# Tile 14: 3x Barrows Unique
add_tile(14, "Obtain 3x Barrows Unique", BARROWS_3X_HOURS, f"1/17.42 per chest, {barrows_kph} chests/hr")

# Tile 15: 1x Ring of the Gods, Treasonous Ring or Tyrannical Ring
# These drop from wilderness bosses (Vet'ion, Venenatis, Callisto) and their demi-boss counterparts
p = 1/716; kph = 50
med = ceil(log(0.5) / log(1 - p))
add_tile(15, "Obtain 1x Wildy Boss Ring", med/kph, f"1/716 from artio, {kph} kph, median {med} kc")

# Tile 16: 3x Glacial Temotli (from Amoxliatl)
p = 1/100; kph = 71
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(16, "Obtain 3x Glacial Temotli", med/kph, f"1/100 from Amoxliatl, {kph} kph, median {med} kc", confidence="high")

# Tile 17: 1x Warped Sceptre (from Warped Terrorbirds)
p = 1/320; kph = 120
med = int(ceil(nbinom.ppf(0.5, 1, p)))
add_tile(17, "Obtain 1x Warped Sceptre", med/kph, f"1/320 from terrorbirds, {kph} kph, median {med} kc", confidence="high")

# Tile 18: 1x Sulphur Blades (from Sulphur Naguas)
p = 1/450; kph = 290 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_sulphur_naguas
med = int(ceil(nbinom.ppf(0.5, 1, p)))
add_tile(18, "Obtain 1x Sulphur Blades", med/kph, f"1/450 from sulphur naguas, {kph} kph, median {med} kc")

# Tile 19: Movement
add_movement(19, "Go back to Tile #14", 14)

# Tile 20: 1x Sarachnis Cudgel (1/384 from Sarachnis, ~40 kph)
p = 1/384; kph = 67
med = ceil(log(0.5) / log(1 - p))
add_tile(20, "Obtain 1x Sarachnis Cudgel", med/kph, f"1/384 from Sarachnis, {kph} kph, median {med} kc")

# Tile 21: 3x Giant Key (from Obor)
p = 1/16; kph = 120
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(21, "Obtain 3x Giant Key", med/kph, f"1/16 from Obor, burning claw spec+desert ammy/house+giantsoul, ~{kph} kph, median {med} kc")

# Tile 22: 5x Steel Ring (Deranged Arch)
p = 1/44; kph = 95
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(22, "Obtain 5x Steel Ring", med/kph, f"1/44 from Deranged arch, ~{kph} kph, median {med} kc")

# Tile 23: 3x Alchemist's Signet (from elder custodian stalkers)
p = 1/62; kph = ELDER_CUSTODIANS_PER_HOUR
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(23, "Obtain 3x Alchemist's Signet", med/kph, f"1/62 from elder custodian stalkers, ~{kph} kph, median {med} kc")

# Tile 24: 5x Giantsoul Amulet (from Giant bosses area?)
p = 1/32; kph = 55 # 1/16 drop rate, 50% contribution from duo https://oldschool.runescape.wiki/w/Royal_Titans#Rewards
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(24, "Obtain 5x Giantsoul Amulet", med/kph, f"1/32 from equal contribution royal titans, ~{kph} kph, median {med} kc")

# Tile 25: 1x Raid unique
add_tile(25, "Obtain 1x Raid Drop (any raid)", RAID_1X_HOURS, "Best via CoX/ToB/ToA team")

# Tile 26: Movement + SIT
add_movement(26, "Advance to Tile #40 (SIT)", 40)

# Tile 27: Movement
add_movement(27, "Advance to Tile #37", 37)

# Tile 28: 1x Beginner Clue Unique
# Beginner clue uniques: Mole slippers, Frog slippers, etc. 
# ~15 beginner clues/hr (pickpocketing HAM members + solving)
# Each clue has ~1/14 chance for a unique? Actually beginner clues give 1 unique roll
# With ~15+ possible uniques, P(any specific) is low but P(ANY unique) might be decent
# Actually beginner caskets can give one of several uniques, I think close to 1/3 chance of any unique
# At 15 clues/hr, median ~1 kill for 1/3 -> like 2 clues = 8 minutes?
# Let's be more conservative: maybe 0.3 hr
add_tile(28, "Obtain 1x Beginner Clue Unique", 0.5, "~15 beginner clues/hr, decent unique chance", confidence="medium")

# Tile 29: 1x Slayer Boss Drop
add_tile(29, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS, "Best via Alchemical Hydra (~1/46 combined unique, 28 kph)", confidence="medium")

# Tile 30: 1x Rev Unique (unique or ancient statuette table)
# Rev orks skulled off-task
p = 1/1000; kph = 110 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_revenants_(Magic_shortbow)
med = ceil(log(0.5) / log(1 - p))
add_tile(30, "Obtain 1x Rev Unique", med/kph, f"Revs ~1/1000 unique from orks, {kph} kph")

# Tile 31: Tempoross (same as tile 2)
add_tile(31, "Obtain Tempoross items", TEMPOROSS_HOURS, "Same as tile 2 - 100 soaked pages fastest")

# Tile 32: 3x Black Mask (1/512 from Cave Horrors, ~200 kph)
p = 1/512; kph = 200
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(32, "Obtain 3x Black Mask", med/kph, f"1/512 from Cave Horrors, {kph} kph, median {med} kc")

# Tile 33: 1x Raid unique
add_tile(33, "Obtain 1x Raid Drop", RAID_1X_HOURS, "Best via CoX/ToB/ToA team")

# Tile 34: 1x Odium Shard (any) - from Crazy Archaeologist, Chaos Fanatic, or Scorpia
# Crazy Archaeologist: 1/256, ~60 kph
p = 1/256; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(34, "Obtain 1x Odium Shard", med/kph, f"1/256 from Crazy Archaeologist, {kph} kph, median {med} kc")

# Tile 35: 1x Egg Sack (from grubby chest)
p = 1/25; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(35, "Obtain 1x orange/blue Egg Sack", med/kph, f"From grubby chest ~1/20, {kph} kph, median {med} kc", confidence="medium")

# Tile 36: 1x Crystal Armour Seed
add_tile(36, "Obtain 1x Crystal Armour Seed", CRYSTAL_SEED_1X_HOURS, "From CG 1/50, ~6/hr")

# Tile 37: 1x Zombie Axe (from armoured zomebies)
p = 1/800; kph = 400
med = ceil(log(0.5) / log(1 - p))
add_tile(37, "Obtain 1x Zombie Axe", med/kph, f"From armoured zombies, {kph} kph, median {med} kc")

# Tile 38: 5x Fresh Crab Shell (from crabs somewhere, similar to claw)
p = 1/8; kph = 150
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(38, "Obtain 5x Fresh Crab Shell", med/kph, f"1/8 from level 23 crabclaw isle crabs, {kph} kph, median {med} kc", confidence="high")

# Tile 39: 1x Hill Giant Club (from Obor, 1/118)
# KPH = time to get new keys from obor with spec/tele tech
p = 1/118; kph = 12
med = int(ceil(nbinom.ppf(0.5, 1, p)))
add_tile(39, "Obtain 1x Hill Giant Club", med/kph, f"1/118 from Obor, {kph} kph, median {med} kc")

# Tile 40: Movement (SIT, go back to 38)
add_movement(40, "Go back to Tile #38 (SIT)", 38)

# Tile 41: 1x Slayer Boss
add_tile(41, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS, "Hydra")

# Tile 42: 1x Elder Chaos Druid Robes piece (from Elder Chaos Druids, 1/1419 each piece, 3 pieces)
p = 3/1419; kph = 200
med = ceil(log(0.5) / log(1 - p))
add_tile(42, "Obtain 1x Elder Chaos Druid Robe piece", med/kph, f"3/1419 combined, {kph} kph, median {med} kc")

# Tile 43: 3x Moons of Peril Unique
add_tile(43, "Obtain 3x Moons of Peril Unique", moons_hours(3), f"~1/19 combined unique, {moons_kph} kph", confidence="high")

# Tile 44: 1x Raid Drop
add_tile(44, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 45: 1x Fedora (from Crazy Archaeologist)
# Fedora: 1/128 from Crazy Archaeologist, ~60 kph
p = 1/128; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(45, "Obtain 1x Fedora", med/kph, f"1/128 from Crazy Archaeologist, {kph} kph")

# Tile 46: Free
add_free(46, "Free Tile - Roll Again")

# Tile 47: 1x Godsword Shard (from any GWD boss, 1/512)
# Shards are 1/512 each, 3 shards total. Any shard = 3/512 = 1/170.7
# ~30 kph at any GWD boss
p = 3/512; kph = 30
med = ceil(log(0.5) / log(1 - p))
add_tile(47, "Obtain 1x Godsword Shard", med/kph, f"3/512 for any shard, {kph} kph at GWD")

# Tile 48: 1x Ice or Fire Elemental Staff Crown
p = 2/150; kph = 55
med = ceil(log(0.5) / log(1 - p))
add_tile(48, "Obtain 1x Elemental Staff Crown", med/kph, f"2/150 for either from equal contribution royal titans, ~{kph} kph, median {med} kc")

# Tile 49: 3x Right Skull Half (S.S. minotaurs)
p = 1/33; kph = 180
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(49, "Obtain 3x Right Skull Half", med/kph, f"1/33 from S.S. minotaurs, ~{kph} kph, median {med} kc")

# Tile 50: Movement
add_movement(50, "Advance to Tile #62", 62)

# Tile 51: 1x Ancient Ceremonial Robes piece (from Nex's minions in GWD)
# Blowpipe blood reavers
p = 1/640; kph = 120
med = ceil(log(0.5) / log(1 - p))
add_tile(51, "Obtain 1x Ancient Ceremonial piece", med/kph, f"1/640 from blood reavers outside nex bank, {kph} kph, median {med} kc", confidence="medium")

# Tile 52: 3x Amulet of the Damned (from Shade catacombs chests)
p = 1/15; kph = 60
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(52, "Obtain 3x Amulet of the Damned", med/kph, f"1/15 from shades silver/red chests chests, {kph} kph, median {med} kc")

# Tile 53: 1x Pharaoh's Sceptre (from Pyramid Plunder)
# 1/75 chance per run at 91+ thieving https://oldschool.runescape.wiki/w/Pharaoh%27s_sceptre#Obtaining
p = 1/75; kph = 8
med = ceil(log(0.5) / log(1 - p))
add_tile(53, "Obtain 1x Pharaoh's Sceptre", med/kph, "~1/75 effective per 91 thieving PP run, 8 runs/hr")

# Tile 54: Movement
add_movement(54, "Go back to Tile #43", 43)

# Tile 55: 30 Molch Pearls (from Aerial Fishing)
# ~8 pearls/hr at Aerial Fishing (rough estimate)
add_tile(55, "Obtain 30 Molch Pearls", 30/8, "~8 pearls/hr from Aerial Fishing", confidence="medium")

# Tile 56: 2x Green/Red/Blue Abyssal Dye
# From GOTR? Or from Abyssal creatures?
p = 3/1200; kph = 30 # 30 reward pulls per hour
med = int(ceil(nbinom.ppf(0.5, 2, p)))
add_tile(56, "Obtain 2x Abyssal Dye", med/kph, "30 pulls per hour, 1/1200 drop for each dye, 3/1200 for any, {kph} permits/hour, median {med} permits", confidence="medium")

# Tile 57: 1x Raid Drop
add_tile(57, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 58: 1x Brine Sabre (from Brine Rat, 1/512, ~100 kph)
p = 1/512; kph = 100
med = ceil(log(0.5) / log(1 - p))
add_tile(58, "Obtain 1x Brine Sabre", med/kph, f"1/512 from Brine Rats, {kph} kph, median {med} kc")

# Tile 59: 1x Hueycoatl Unique
add_tile(59, "Obtain 1x Hueycoatl Unique", huey_hours(1), "Includes hides ~1/70 20 kph trio", confidence="medium")

# Tile 60: 1x GWD Drop
add_tile(60, "Obtain 1x GWD Drop", GWD_1X_HOURS, f"~1/103 combined at Zilyana, {gwd_kph} kph")

# Tile 61: 1x Slayer Boss
add_tile(61, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 62: 3x DK Ring
add_tile(62, "Obtain 3x DK Ring", DK_3X_HOURS)

# Tile 63: 4x Barrows
add_tile(63, "Obtain 4x Barrows Unique", BARROWS_4X_HOURS)

# Tile 64: 1x Slayer Boss
add_tile(64, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 65: 1x Crystal Armour Seed
add_tile(65, "Obtain 1x Crystal Armour Seed", CRYSTAL_SEED_1X_HOURS)

# Tile 66: 1x Raid Drop
add_tile(66, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 67: 3x Venator Shard (muspah)
p = 1/100; kph = 25
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(67, "Obtain 3x Venator Shard", med/kph, f"1/100 from muspah, {kph} kph, median {med} kc")

# Tile 68: Movement
add_movement(68, "Advance to Tile #76", 76)

# Tile 69: 3x Ecumenical Key
add_tile(69, "Obtain 3x Ecumenical Key", ECUMENICAL_3X_HOURS, "1/40, wildy GWD, kill imps/goblins")

# Tile 70: 1x Dragon Pickaxe (from KBD, Chaos Ele, Venenatis, Vet'ion, Callisto, or KQ)
p = 1/256; kph = 48
med = ceil(log(0.5) / log(1 - p))
add_tile(70, "Obtain 1x Dragon Pickaxe", med/kph, f"1/256 from chaos ele, {kph} kph, median {med} kc")

# Tile 71: 1x Zenyte Shard (from Demonic Gorillas, 1/300, ~60 kph)
p = 1/300; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(71, "Obtain 1x Zenyte Shard", med/kph, f"1/300 from Demonic Gorillas, {kph} kph, median {med} kc")

# Tile 72: 1x Champion Scroll
add_tile(72, "Obtain 1x Champion Scroll", CHAMPION_SCROLL_HOURS, "1/5000, cannoning goblins ~500/hr")

# Tile 73: 1x Slayer Boss
add_tile(73, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 74: 5x Flippers (from Mogres, 1/64)
p = 1/64; kph = 40
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(74, "Obtain 5x Flippers", med/kph, f"1/64 from Mogres, {kph} kph, median {med} kc")

# Tile 75: 1x Teleport Anchoring Scroll (Zombie chest)
p = 1/275; kph = 200
med = ceil(log(0.5) / log(1 - p))
add_tile(75, "Obtain 1x Teleport Anchoring Scroll", med / kph, "Looting the zombie pirate's locker, 1/275, 200 kph")

# Tile 76: 1x Hueycoatl Unique
add_tile(76, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 77: 1x Raid Drop
add_tile(77, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 78: 3x Barrows
add_tile(78, "Obtain 3x Barrows Unique", BARROWS_3X_HOURS)

# Tile 79: 1x TzHaar weapon/armour
p = 1/300; kph = 300
med = ceil(log(0.5) / log(1 - p))
add_tile(79, "Obtain 1x TzHaar Weapon/Armour", med/kph, f"~1/300 combined obsidian, {kph} kph barraging", confidence="medium")

# Tile 80: 1x Shark Paint (1/36 upon completing a port task)
p = 1/36; kph = 20
med = ceil(log(0.5) / log(1 - p))
add_tile(80, "Obtain 1x Shark Paint", med/kph, "1/36 from port tasks, assuming a very slow 20 per hour")

# Tile 81: 1x Cache of Runes
p = 3/27; kph = 3
med = ceil(log(0.5) / log(1 - p))
add_tile(81, "Obtain 1x Cache of Runes", med/kph, "3/27 from ToA chest, assuming 20 minute 150's")

# Tile 82: 1x Tertiary Drop from Zalcano
# Zalcano tertiary: Crystal tool seed (1/200) split 3 ways in a trio, 1/600
# Smolcano pet (1/2250), 
# Zalcano shard (~1/1000).
# ~1/540 in a trio
# ~30 kph
p = 1/540; kph = 30
med = ceil(log(0.5) / log(1 - p))
add_tile(82, "Obtain 1x Zalcano Tertiary", med/kph, f"Combined ~1/540 in an efficient trio, {kph} kph, median {med} kc")

# Tile 83: 1x DT2 Boss unique + secondary
add_tile(83, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS, "Vardorvis best with chromium ingot 1/150")

# Tile 84: 1x Slayer Boss
add_tile(84, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 85: Movement
add_movement(85, "Advance to Tile #92", 92)

# Tile 86: 1x Any Big Fish (1/1000 for bass)
p = 1/1000; kph = 120
med = ceil(log(0.5) / log(1 - p))
add_tile(86, "Obtain 1x Any Big Fish", med/kph, f"Fishing bass at 99, {kph} per hour, {med} fish median", confidence="medium")

# Tile 87: 1x Rev Unique
p = 1/1000; kph = 110 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_revenants_(Magic_shortbow)
med = ceil(log(0.5) / log(1 - p))
add_tile(87, "Obtain 1x Rev Unique", med/kph, f"Same as tile 30, Revs ~1/1000 unique from orks, {kph} kph")

# Tile 88: 3x Barrows
add_tile(88, "Obtain 3x Barrows Unique", BARROWS_3X_HOURS)

# Tile 89: 1x Raid Drop
add_tile(89, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 90: 1x Crystal or Enhanced Crystal Weapon Seed (NO LMS)
# Crystal weapon seed: 1/50 from CG (same as armour seed)
# Enhanced crystal weapon seed: 1/400 from CG
# Combined: 1/50 + 1/400 = 9/400 = 1/44.4
# 6 CG/hr
p = 9/400; kph = 6
med = ceil(log(0.5) / log(1 - p))
add_tile(90, "Obtain 1x Crystal/Enhanced Weapon Seed", med/kph, f"Combined ~1/44 from CG, 6 kph")

# Tile 91: 1x Hueycoatl Unique
add_tile(91, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 92: Wintertodt
add_tile(92, "Obtain Wintertodt items", WINTERTODT_HOURS, "100 burnt pages fastest")

# Tile 93: 3x Whip or 1x Unsired
add_tile(93, "Obtain 3x Whip or 1x Unsired", WHIP_OR_UNSIRED_3X_HOURS, "1x Unsired from Sire fastest (~1/100, 28 kph)")

# Tile 94: 3x Moons of Peril
add_tile(94, "Obtain 3x Moons of Peril Unique", moons_hours(3))

# Tile 95: 1x Frozen Cache (muspah)
p = 1/72; kph = 25
med = ceil(log(0.5) / log(1 - p))
add_tile(95, "Obtain 1x Frozen Cache", med/kph, f"From muspah, {kph} kph, {med} median kc")

# Tile 96: 1x Slayer Boss
add_tile(96, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 97: 1x Raid Drop
add_tile(97, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 98: 1x Elder Chaos Druid Robe piece
p = 3/1419; kph = 200
med = ceil(log(0.5) / log(1 - p))
add_tile(98, "Obtain 1x Elder Chaos Druid Robe piece", med/kph, f"Same as tile 42, 3/1419 combined, {kph} kph, median {med} kc")

# Tile 99: Movement
add_movement(99, "Go back to Tile #94", 94)

# Tile 100: 1x Zombie Helmet (from Armoured Zombies)
p = 1/600; kph = 400
med = ceil(log(0.5) / log(1 - p))
add_tile(100, "Obtain 1x Broken Zombie helmet", med/kph, f"From Zemouregal's fort armoured zombies, {kph} kph, median {med} kc")

# Tile 101: 3x Easy Clue Uniques
p = 247/1080; kph = 10 # https://oldschool.runescape.wiki/w/Reward_casket_(easy)
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(101, "Obtain 3x Easy Clue Uniques", med/kph, f"10 easy clues/hr, 247/1080 chance for a unique from each")

# Tile 102: 1x Slayer Boss
add_tile(102, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 103: 1x Doom of Mokhaiotl Unique
add_tile(103, "Obtain 1x Doom Unique", doom_hours(1), "~1/50 for any unique by wave 8, 6 kph")

# Tile 104: 3x Moons of Peril
add_tile(104, "Obtain 3x Moons of Peril Unique", moons_hours(3))

# Tile 105: Amoxliatl Speed-Trialist (sub 1 min kill)
add_tile(105, "Complete Amoxliatl Speed-Trialist", 0.1, "Marked 'Practically Free' - just need sub-1min kill", confidence="high")

# Tile 106: 3x DK Ring
add_tile(106, "Obtain 3x DK Ring", DK_3X_HOURS)

# Tile 107: 1x Colored Egg Sack (from grubby chest)
p = 1/25; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(107, "Obtain 1x orange/blue Egg Sack", med/kph, f"Same as tile 35, from grubby chest ~1/20, {kph} kph, median {med} kc")

# Tile 108: 1x Gnome Restaurant unique (Gnome Scarf/Goggles/Mint Cake)
# https://oldschool.runescape.wiki/w/Money_making_guide/Delivering_food_in_Gnome_Restaurant
add_tile(108, "Obtain 1x Gnome Restaurant unique", 2.0, "0.6 each of scarf/goggles/mint cakes expected in 1 hour of delivery, https://oldschool.runescape.wiki/w/Money_making_guide/Delivering_food_in_Gnome_Restaurant")

# Tile 109: 1x Slayer Boss
add_tile(109, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 110: Movement
add_movement(110, "Advance to Tile #114", 114)

# Tile 111: 1x Crystal Armour Seed
add_tile(111, "Obtain 1x Crystal Armour Seed", CRYSTAL_SEED_1X_HOURS)

# Tile 112: 1x Zombie Axe
p = 1/800; kph = 400
med = ceil(log(0.5) / log(1 - p))
add_tile(112, "Obtain 1x Zombie Axe", med/kph, f"Same as tile 37, from armoured zombies, {kph} kph, median {med} kc")

# Tile 113: 4x Barrows
add_tile(113, "Obtain 4x Barrows Unique", BARROWS_4X_HOURS)

# Tile 114: 1x Raid Drop
add_tile(114, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 115: 1x Crawling Hand (from Crawling Hands)
p = 1/500; kph = 300
med = ceil(log(0.5) / log(1 - p))
add_tile(115, "Obtain 1x Crawling Hand", med/kph, f"1/300 From Crawling Hand monsters, {kph} kph, {med} kc")

# Tile 116: Free
add_free(116, "Free Tile - Roll Again")

# Tile 117: 1x Zulrah Unique
add_tile(117, "Obtain 1x Zulrah Unique", ZULRAH_1X_HOURS, "~1/155 combined with mutagens, 35 kph")

# Tile 118: 2x Antler Guard
p = 1/650; kph = ELDER_CUSTODIANS_PER_HOUR
med = int(ceil(nbinom.ppf(0.5, 2, p)))
add_tile(118, "Obtain 2x Antler Guard", med/kph, f"Similar to tile 12, cannoning Elder custodian stalkers, ~{kph} per hour, median {med} kc", confidence="high")

# Tile 119: 1x Slayer Boss
add_tile(119, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 120: 1x Hueycoatl
add_tile(120, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 121: 4x Barrows
add_tile(121, "Obtain 4x Barrows Unique", BARROWS_4X_HOURS)

# Tile 122: 1x Chewed Bones (from Mithril Dragons, 1/42)
p = 3/128; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(122, "Obtain 1x Chewed Bones", med/kph, f"3/128 from Mithril Dragons, {kph} kph, median {med} kc")

# Tile 123: Free (+1 Skip)
add_free(123, "Gain +1 SKIP - Roll Again")

# Tile 124: 1x Forgotten Lockbox (Yama)
p = 1/33; kph = 8 # solo
med = ceil(log(0.5) / log(1 - p))
add_tile(124, "Obtain 1x Forgotten Lockbox", med/kph, f"1/33 from solo yama, {kph} kph, median {med} kc, duo rate should be similar")

# Tile 125: 1x Raid Drop
add_tile(125, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 126: 3x Moons of Peril
add_tile(126, "Obtain 3x Moons of Peril Unique", moons_hours(3))

# Tile 127: 1x Slayer Boss
add_tile(127, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 128: 1x Skull of Vet'ion/Claws of Callisto/Fang of Venenatis
p = 1/618; kph = 50
med = ceil(log(0.5) / log(1 - p))
add_tile(128, "Obtain 1x Wildy Boss Weapon upgrade", med/kph, f"1/618 from artio, {kph} kph")

# Tile 129: 1x Earthbound Tecpatl (newer content)
p = 1/400; kph = 200
med = ceil(log(0.5) / log(1 - p))
add_tile(129, "Obtain 1x Earthbound Tecpatl", 2.0, "1/400 from earthen nagua, 200 kph?", confidence="medium")

# Tile 130: 3x Dragon Boots (from Spiritual Mages, 1/128)
p = 1/128; kph = 180  # blowpiping nex spiritual mages
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(130, "Obtain 3x Dragon Boots", med/kph, f"1/128 from Nex Spiritual Mages, {kph} kph, median {med} kc")

# Tile 131: Movement
add_movement(131, "Go back to Tile #126", 126)

# Tile 132: Movement
add_movement(132, "Advance to Tile #140", 140)

# Tile 133: 1x Doom Unique
add_tile(133, "Obtain 1x Doom Unique", doom_hours(1))

# Tile 134: 1x Slayer Boss
add_tile(134, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 135: Movement
add_movement(135, "Go back to Tile #121", 121)

# Tile 136: 1x Medium clue boots (Ranger/Climbing(g)/Holy Sandals/Spiked Manacles/Wizard)
p = 5/283.6; kph = 10 # 10 clues/hour from eclectics
med = ceil(log(0.5) / log(1 - p))
add_tile(136, "Obtain 1x Med Clue Boots", med/kph, f"5/283.6 combined from med caskets, {kph} clues/hr, median {med} caskets")

# Tile 137: 1x Raid Drop
add_tile(137, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 138: 1x Granite Maul (from Gargoyles, 1/256, ~200 kph)
p = 1/256; kph = 200
med = ceil(log(0.5) / log(1 - p))
add_tile(138, "Obtain 1x Granite Maul", med/kph, f"1/256 from Gargoyles, {kph} kph, median {med} kc")

# Tile 139: 3x Ecumenical Key
add_tile(139, "Obtain 3x Ecumenical Key", ECUMENICAL_3X_HOURS)

# Tile 140: Tempoross
add_tile(140, "Obtain Tempoross items", TEMPOROSS_HOURS)

# Tile 141: 2x Slayer Boss
add_tile(141, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 142: 1x Raid Drop
add_tile(142, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 143: 1x GWD Drop
add_tile(143, "Obtain 1x GWD Drop", GWD_1X_HOURS)

# Tile 144: 5x Medium Clue Uniques
# ~10 med clues/hr, ~3/10 for a unique from each casket
p = 3/10; kph = 8
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(144, "Obtain 5x Medium Clue Uniques", med/kph, f"~3/10 any unique per casket, {kph} clues/hr", confidence="medium")

# Tile 145: Movement
add_movement(145, "Advance to Tile #160", 160)

# Tile 146: 2x Slayer Boss
add_tile(146, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 147: 1x Vorkath Unique
add_tile(147, "Obtain 1x Vorkath Unique", VORKATH_UNIQUE_HOURS, "Head at 1/50 makes this fast, 25 kph")

# Tile 148: Free
add_free(148, "Free Tile - Roll Again")

# Tile 149: 1x Raid
add_tile(149, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 150: 1x Dragon 2h Sword (from Chaos Elemental/KBD? Actually from Chaos Elemental 1/128, or rare drop table)
# Dragon 2h sword is an RDT item or from specific bosses
# Chaos Elemental: 1/128, ~48 kph
p = 1/64; kph = 48
med = ceil(log(0.5) / log(1 - p))
add_tile(150, "Obtain 1x Dragon 2h Sword", med/kph, f"1/128 from Chaos Elemental, {kph} kph")

# Tile 151: 1x Enhanced Crystal Teleport Seed (thieving)
add_tile(151, "Obtain 1x Enhanced Crystal Teleport Seed", 2.0, "~1 per hour with thieving outfit, so 2 hours pickpocketing to see a drop https://oldschool.runescape.wiki/w/Money_making_guide/Pickpocketing_elves")

# Tile 152: 1x Zulrah Unique
add_tile(152, "Obtain 1x Zulrah Unique", ZULRAH_1X_HOURS)

# Tile 153: 1x DT2 Boss Drop
add_tile(153, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 154: 2x Slayer Boss
add_tile(154, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 155: 1x GWD Drop
add_tile(155, "Obtain 1x GWD Drop", GWD_1X_HOURS)

# Tile 156: 4x Barrows
add_tile(156, "Obtain 4x Barrows Unique", BARROWS_4X_HOURS)

# Tile 157: Movement
add_movement(157, "Go back to Tile #146", 146)

# Tile 158: 1x Raid
add_tile(158, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 159: 1x Malediction Shard (from Crazy Arch/Chaos Fanatic/Scorpia)
# Same rates as Odium Shard
p = 1/256; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(159, "Obtain 1x Malediction Shard", med/kph, f"1/256 from Crazy Archaeologist, {kph} kph")

# Tile 160: 1x Raid
add_tile(160, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 161: 25x Marks of Grace
add_tile(161, "Obtain 25x Mark of Grace", MARKS_25_HOURS)

# Tile 162: 2x Slayer Boss
add_tile(162, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 163: 1x Elder Chaos Druid Robe
add_tile(163, "Obtain 1x Elder Chaos Druid Robe", 0.16)

# Tile 164: Movement
add_movement(164, "Advance to Tile #169", 169)

# Tile 165: Movement
add_movement(165, "Advance to Tile #172", 172)

# Tile 166: 1x Tormented Synapse (Tormented Demons)
p = 1/500; kph = 55 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_Tormented_Demons
med = ceil(log(0.5) / log(1 - p))
add_tile(166, "Obtain 1x Tormented Synapse", med/kph, f"1/500 from TD's, {kph} kph, median {med} kc")

# Tile 167: 1x Granite Maul
add_tile(167, "Obtain 1x Granite Maul", 0.89, "1/256 from Gargoyles, 200 kph")

# Tile 168: 1x Bottom + 1x Top of Sceptre (Stronghold of security)
add_tile(168, "Obtain Sceptre pieces (Runed Sceptre)", 1.0, "1/33 drops from easy low HP monsters")

# Tile 169: 4x Moons of Peril
add_tile(169, "Obtain 4x Moons of Peril Unique", moons_hours(4))

# Tile 170: 1x Raid
add_tile(170, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 171: 3x DK Ring
add_tile(171, "Obtain 3x DK Ring", DK_3X_HOURS)

# Tile 172: 3x Whip or 1x Unsired
add_tile(172, "Obtain 3x Whip or 1x Unsired", WHIP_OR_UNSIRED_3X_HOURS)

# Tile 173: 1x Crystal Armour Seed
add_tile(173, "Obtain 1x Crystal Armour Seed", CRYSTAL_SEED_1X_HOURS)

# Tile 174: Movement
add_movement(174, "Go back to Tile #161", 161)

# Tile 175: 1x Abyssal Dye
p = 3/1200; kph = 30 # 30 reward pulls per hour
med = int(ceil(nbinom.ppf(0.5, 1, p)))
add_tile(56, "Obtain 1x Abyssal Dye", med/kph, "Similar to tile 56, 30 pulls per hour, 1/1200 drop for each dye, 3/1200 per pull, {kph} permits/hour, median {med} permits", confidence="medium")

# Tile 176: 1x Doom Unique
add_tile(176, "Obtain 1x Doom Unique", doom_hours(1))

# Tile 177: 1x Raid
add_tile(177, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 178: 1x Hueycoatl
add_tile(178, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 179: 1x Bloody Notes (Shades of Mort'ton chests)
# 60 chests per hour, ~1/105 from gold chests
p = 1/105; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(179, "Obtain 1x Bloody Notes", med/kph, f"1/105 from gold catacombs chests, 60 kph")

# Tile 180: 1x Zenyte Shard
p = 1/300; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(180, "Obtain 1x Zenyte Shard", med/kph, f"Same as tile 71, 1/300 from Demonic Gorillas, {kph} kph, median {med} kc")

# Tile 181: 1x Burning Claw (from TD's)
p = 1/501; kph = 55 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_Tormented_Demons
med = ceil(log(0.5) / log(1 - p))
add_tile(181, "Obtain 1x burning claw", med/kph, f"1/501 from TD's, {kph} kph, median {med} kc")

# Tile 182: Movement
add_movement(182, "Go back to Tile #171", 171)

# Tile 183: 2x Slayer Boss
add_tile(183, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 184: 1x Champion Scroll
add_tile(184, "Obtain 1x Champion Scroll", CHAMPION_SCROLL_HOURS)

# Tile 185: 1x Rev Unique
p = 1/1000; kph = 110 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_revenants_(Magic_shortbow)
med = ceil(log(0.5) / log(1 - p))
add_tile(185, "Obtain 1x Rev Unique", med/kph, f"Same as tile 30, Revs ~1/1000 unique from orks, {kph} kph")

# Tile 186: 1x DT2 Boss Drop
add_tile(186, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 187: 1x Raid
add_tile(187, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 188: 1x Slayer Boss
add_tile(188, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 189: Movement
add_movement(189, "Advance to Tile #196", 196)

# Tile 190: Wintertodt
add_tile(190, "Obtain Wintertodt items", WINTERTODT_HOURS)

# Tile 191: 3x Venator Shard (muspah)
p = 1/100; kph = 25
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(191, "Obtain 3x Venator Shard", med/kph, f"same as tile 67, 1/100 from muspah, {kph} kph, median {med} kc")

# Tile 192: 2x Slayer Boss
add_tile(192, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 193: 1x Raid
add_tile(193, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 194: 1x Elemental Staff Crown
p = 2/150; kph = 55
med = ceil(log(0.5) / log(1 - p))
add_tile(194, "Obtain 1x Elemental Staff Crown", med/kph, f"Same as tile 48, 2/150 for either from equal contribution royal titans, ~{kph} kph, median {med} kc")

# Tile 195: 1x Zulrah Unique
add_tile(195, "Obtain 1x Zulrah Unique", ZULRAH_1X_HOURS)

# Tile 196: 1x Dragon 2h Sword
p = 1/64; kph = 48
med = ceil(log(0.5) / log(1 - p))
add_tile(196, "Obtain 1x Dragon 2h Sword", med/kph, f"Same as tile 150, 1/128 from Chaos Elemental, {kph} kph")

# Tile 197: Free
add_free(197, "Free Tile - Roll Again")

# Tile 198: Movement
add_movement(198, "Go back to Tile #187", 187)

# Tile 199: 1x Slayer Boss
add_tile(199, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 200: 1x Raid
add_tile(200, "Obtain 1x Raid Drop", RAID_1X_HOURS)

# Tile 201: Movement
add_movement(201, "Advance to Tile #215", 215)

# Tile 202: 1x Cache of Runes
p = 3/27; kph = 3
med = ceil(log(0.5) / log(1 - p))
add_tile(202, "Obtain 1x Cache of Runes", med/kph, "Same as tile 81, 3/27 from ToA chest, assuming 20 minute 150's")

# Tile 203: 2x Slayer Boss
add_tile(203, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 204: 3x Silver/Golden Coffin Locks (from Shade catacombs)
p = 1/60; kph = 60
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(204, "Obtain 3x Coffin Locks", 3.0, "From Shade catacombs, ~1/60 from golden/silver chests")

# Tile 205: 1x Ballista Component (from Demonic Gorillas, 1/500ish for any component)
# Ballista spring, Ballista frame, Ballista limbs, Monkey tail
p = 1/180; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(205, "Obtain 1x Ballista Component", med/kph, f"~1/180 combined from DGs, {kph} kph")

# Tile 206: 1x GWD Drop
add_tile(206, "Obtain 1x GWD Drop", GWD_1X_HOURS)

# Tile 207: 1x Crystal Armour Seed
add_tile(207, "Obtain 1x Crystal Armour Seed", CRYSTAL_SEED_1X_HOURS)

# Tile 208: 2x Raid Drops
add_tile(208, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 209: 1x Doom Unique
add_tile(209, "Obtain 1x Doom Unique", doom_hours(1))

# Tile 210: 5x Medium Clue Uniques
p = 3/10; kph = 8
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(210, "Obtain 5x Medium Clue Uniques", med/kph, f"Same as tile 144, ~3/10 any unique per casket, {kph} clues/hr")

# Tile 211: 1x Odium Shard
p = 1/256; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(211, "Obtain 1x Odium Shard", med/kph, f"Same as tile 34, 1/256 from Crazy Archaeologist, {kph} kph, median {med} kc")

# Tile 212: 1x Slayer Boss
add_tile(212, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 213: Movement
add_movement(213, "Go back to Tile #204", 204)

# Tile 214: 3x Moons of Peril
add_tile(214, "Obtain 3x Moons of Peril Unique", moons_hours(3))

# Tile 215: 2x Raid Drops
add_tile(215, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 216: 1x Frozen Cache
p = 1/72; kph = 25
med = ceil(log(0.5) / log(1 - p))
add_tile(216, "Obtain 1x Frozen Cache", med/kph, f"Same as tile 95, From muspah, {kph} kph, {med} median kc")

# Tile 217: 1x Slayer Boss
add_tile(217, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 218: 1x Sarachnis Cudgel
p = 1/384; kph = 67
med = ceil(log(0.5) / log(1 - p))
add_tile(218, "Obtain 1x Sarachnis Cudgel", med/kph, "Same as tile 20, 1/384, 67 kph")

# Tile 219: 1x Wildy boss wep upgrade
p = 1/618; kph = 50
med = ceil(log(0.5) / log(1 - p))
add_tile(219, "Obtain 1x Wildy Boss Weapon upgrade", med/kph, f"Same as tile 128, 1/618 from artio, {kph} kph")

# Tile 220: 2x Ancient Ceremonial piece
p = 1/640; kph = 120
med = int(ceil(nbinom.ppf(0.5, 2, p)))
add_tile(220, "Obtain 2x Ancient Ceremonial piece", med/kph, f"Similar to tile 51, 1/640 from blood reavers outside nex bank, {kph} kph, median {med} kc", confidence="medium")

# Tile 221: 1x Shaman Mask (Ogress shamans)
p = 1/1200; kph = 120
med = ceil(log(0.5) / log(1 - p))
add_tile(221, "Obtain 1x Shaman Mask", med/kph, "1/1200 from ogress shamans/warriors, 120 kph")

# Tile 222: 1x Dragon Axe (from DKs or Wintertodt)
# DKs: Dragon axe 1/128 from Dagannoth Rex? Actually it's from all DK kings
# Actually Dragon Axe is from Dagannoth Kings (all three can drop it)
# Rex: 1/128, Supreme: 1/128, Prime: 1/128 -> combined 3/128 = 1/42.67
# 45 kills/hr total -> median 29/45 = 0.64 hr
# Or from Wintertodt: 1/10000 per crate - way slower
p = 3/128; kph = 45
med = ceil(log(0.5) / log(1 - p))
add_tile(222, "Obtain 1x Dragon Axe", med/kph, f"1/42.7 combined from DKs, {kph} kph")

# Tile 223: SIT (Lose -1 SKIP)
add_free(223, "Lose -1 SKIP - Roll Again (SIT)")

# Tile 224: 1x Nightmare Unique
# Using phosani's numbers
p = 1/113; kph = 9 # https://oldschool.runescape.wiki/w/Phosani%27s_Nightmare#Uniques
med = ceil(log(0.5) / log(1 - p))
add_tile(224, "Obtain 1x Nightmare Unique", 11.0, "Phosani's ~1/113 combined unique chance, 9 kph")

# Tile 225: 1x GWD Drop
add_tile(225, "Obtain 1x GWD Drop", GWD_1X_HOURS)

# Tile 226: Gnome Restaurant
add_tile(226, "Obtain 1x Gnome Restaurant unique", 2.0, "Same as tile 108, 0.6 each of scarf/goggles/mint cakes expected in 1 hour of delivery, https://oldschool.runescape.wiki/w/Money_making_guide/Delivering_food_in_Gnome_Restaurant")

# Tile 227: 1x SRA Piece
# Vard is the same EHB since the duke changes with a lower axe pc drop rate
# Whisperer is ~half the EHB rate but the piece is more than half as rare (1/512)
# Leviathan piece is rarer and he's slower to kill than duke
p = 1/720; kph = 40 # Duke numbers
med = ceil(log(0.5) / log(1 - p))
add_tile(227, "Obtain 1x SRA Piece", med/kph, f"Uses duke numbers, 40 kph 1/720")

# Tile 228: 1x TzHaar weapon/armour
p = 1/300; kph = 300
med = ceil(log(0.5) / log(1 - p))
add_tile(228, "Obtain 1x TzHaar Weapon/Armour", med/kph, f"Same as tile 79, ~1/300 combined obsidian, {kph} kph barraging", confidence="medium")

# Tile 229: 2x Raid Drops
add_tile(229, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 230: 1x Elven Signet (Crystal implings)
p = 1/128; kph = 12 # Source for 12 imps per hour: https://www.youtube.com/watch?v=luJwoTbBH-o
med = ceil(log(0.5) / log(1 - p))
add_tile(230, "Obtain 1x Elven Signet", med/kph, "1/128, 12 imps per hour, might not all be able to do at the same time due to long respawn+world hopping")

# Tile 231: 10x Fire Capes
add_tile(231, "Obtain 10x Fire Capes", 5.8, "~35 min per run")

# Tile 232: Free
add_free(232, "Free Tile - Roll Again")

# Tile 233: 1x Slayer Boss
add_tile(233, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 234: 1x Chewed Bones
p = 3/128; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(234, "Obtain 1x Chewed Bones", med/kph, f"Same as tile 122, 3/128 from Mithril Dragons, {kph} kph, median {med} kc")

# Tile 235: Movement
add_movement(235, "Go back to Tile #228", 228)

# Tile 236: 1x Hueycoatl
add_tile(236, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 237: 1x DT2 Boss Drop
add_tile(237, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 238: Movement (SIT)
add_movement(238, "Go back to Tile #223 (SIT)", 223)

# Tile 239: 1x Tormented Synapse
p = 1/500; kph = 55 # https://oldschool.runescape.wiki/w/Money_making_guide/Killing_Tormented_Demons
med = ceil(log(0.5) / log(1 - p))
add_tile(239, "Obtain 1x Tormented Synapse", med/kph, f"Same as tile 166, 1/500 from TD's, {kph} kph, median {med} kc")

# Tile 240: 2x Raid Drops
add_tile(240, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 241: 1x Blood Shard (Pickpocketing)
# I believe killing vyrewatch ends up being ~18 hours on rate
p = 1/5000; kph = 720
med = ceil(log(0.5) / log(1 - p))
add_tile(241, "Obtain 1x Blood Shard", med/kph, f"1/5000 from pickpocketing vyres, {kph} kph, median {med} kc")

# Tile 242: Movement
add_movement(242, "Advance to Tile #253", 253)

# Tile 243: 1x Sigil/Holy Elixir/Spirit Shield (from Corp)
# Corp: Holy Elixir 1/171, Spirit Shield 1/64, Sigils (spectral 1/1365, arcane 1/1365, elysian 1/4095)
# Spirit Shield 1/64 is most common -> combined with elixir: ~1/44
p = 1/44; kph = 7
med = ceil(log(0.5) / log(1 - p))
add_tile(243, "Obtain 1x Corp Drop", med/kph, f"Spirit Shield 1/64, Elixir 1/171 combined ~1/44, {kph} kph")

# Tile 244: 5x Scurrius Spine
p = 1/33; kph = 40
med = int(ceil(nbinom.ppf(0.5, 5, p)))
add_tile(244, "Obtain 5x Scurrius' Spine", med/kph)

# Tile 245: 1x Oathplate Piece/Soulflame Horn/Pet (Yama)
# 5/600 for oath/horn, 0.24/600 for pet
p = (5.24)/600; kph = 8
med = ceil(log(0.5) / log(1 - p))
add_tile(245, "Obtain 1x Oathplate/Soulflame/Pet", med/kph, "5.4/600 for pet or oath or horn, 8 kph solo (similar rate duo)")

# Tile 246: 3x Easy Clue Uniques
p = 247/1080; kph = 10 # https://oldschool.runescape.wiki/w/Reward_casket_(easy)
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(246, "Obtain 3x Easy Clue Uniques", med/kph, f"Same as tile 101, 10 easy clues/hr, 247/1080 chance for a unique from each")

# Tile 247: 1x Slayer Boss
add_tile(247, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 248: 2x Raid Drops
add_tile(248, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 249: 1x DT2 Boss Drop
add_tile(249, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 250: 1x Champion Scroll
add_tile(250, "Obtain 1x Champion Scroll", CHAMPION_SCROLL_HOURS)

# Tile 251: Movement
add_movement(251, "Go back to Tile #244", 244)

# Tile 252: 1x Bryophyta's Essence (from Bryophyta, 1/118 but need mossy key first)
# 1/16 for a key
# 1/1888 combined
p = 1/1888; kph = 80  # Using burning claws on bryophyta, similar to Obor https://oldschool.runescape.wiki/w/Giant_key, "players can kill Obor 120+ times per hour when using burning claws, giving approximately 8 keys per hour."
med = ceil(log(0.5) / log(1 - p))
add_tile(252, "1x Bryophyta's Essence", med/kph, f"1/16 for a key, essence 1/118 from the chest, ~{kph} burning claw speccing with tele/pool, median {med} kc")

# Tile 253: 1x Frozen Cache
p = 1/72; kph = 25
med = ceil(log(0.5) / log(1 - p))
add_tile(253, "Obtain 1x Frozen Cache", med/kph, f"Same as tile 95, From muspah, {kph} kph, {med} median kc")

# Tile 254: 1x Zulrah Unique
add_tile(254, "Obtain 1x Zulrah Unique", ZULRAH_1X_HOURS)

# Tile 255: 1x Dragon Pickaxe
p = 1/358; kph = 45
med = ceil(log(0.5) / log(1 - p))
add_tile(255, "Obtain 1x Dragon Pickaxe", med/kph, f"Same as tile 70, 1/256 from Calvar'ion, {kph} kph, median {med} kc")

# Tile 256: 1x Echo Crystal (Colo)
add_tile(256, "Obtain 1x Echo Crystal", 4.0, "~0.25/hour completing wave 12 quickly, https://oldschool.runescape.wiki/w/Money_making_guide/Completing_the_Fortis_Colosseum_(Wave_12)", confidence="medium")

# Tile 257: 2x Raid Drops
add_tile(257, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 258: 1x Inky Paint (from krakens)
p = 1/1500; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(258, "Obtain 1x Inky Paint", 2.0, "1/1500 from Vampyre kraken, assuming 60 kph, median {med} kc")

# Tile 259: 5x Crystal Grail (vorpal rabbit)
add_tile(259, "Obtain 5x Crystal Grail", 2.5, "assuming 30 minute rabbit kills")

# Tile 260: 1x Slayer Boss
add_tile(260, "Obtain 1x Slayer Boss Drop", SLAYER_BOSS_1X_HOURS)

# Tile 261: 1x Zenyte Shard
p = 1/300; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(261, "Obtain 1x Zenyte Shard", med/kph, f"Same as tile 71, 1/300 from Demonic Gorillas, {kph} kph, median {med} kc")

# Tile 262: 2x Crystal Armour Seed
p = 1/50; kph = 6
med = int(ceil(nbinom.ppf(0.5, 2, p)))
add_tile(262, "Obtain 2x Crystal Armour Seed", med/kph, f"From CG, median {med} completions")

# Tile 263: 2x Raid Drops
add_tile(263, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 264: 1x Malediction Shard
# Same rates as Odium Shard
p = 1/256; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(264, "Obtain 1x Malediction Shard", med/kph, f"Same as tile 159, 1/256 from Crazy Archaeologist, {kph} kph")

# Tile 265: Free
add_free(265, "Free Tile - Roll Again")

# Tile 266: 1x Doom Unique
add_tile(266, "Obtain 1x Doom Unique", doom_hours(1))

# Tile 267: Movement
add_movement(267, "Advance to Tile #278", 278)

# Tile 268: 2x Slayer Boss
add_tile(268, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 269: 1x Zalcano Tertiary
p = 1/540; kph = 30
med = ceil(log(0.5) / log(1 - p))
add_tile(269, "Obtain 1x Zalcano Tertiary", med/kph, f"Same as tile 82, Combined ~1/540 in an efficient trio, {kph} kph, median {med} kc")

# Tile 270: 2x Raid Drops
add_tile(270, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 271: 1x Oathplate/Soulflame/Pet
p = (5.24)/600; kph = 8
med = ceil(log(0.5) / log(1 - p))
add_tile(271, "Obtain 1x Oathplate/Soulflame/Pet", med/kph, "Same as tile 245, 5.4/600 for pet or oath or horn, 8 kph solo (similar rate duo)")

# Tile 272: 1x Big Fish
p = 1/1000; kph = 120
med = ceil(log(0.5) / log(1 - p))
add_tile(272, "Obtain 1x Any Big Fish", med/kph, f"Same as tile 86, Fishing bass at 99, {kph} per hour, {med} fish median", confidence="medium")

# Tile 273: 1x Holy/Sang/Twisted Kit (HMT/CMs)
# HMT is much faster but I don't think it's realistic for most of the team, me included
p = 1/75; kph = 3
med = ceil(log(0.5) / log(1 - p))
add_tile(273, "Obtain 1x Raid Kit", med/kph, f"Assuming 3CM's/hour, {med} kc on average. HMT would be faster (14 hours, 5/300 for either kit, 3 EHB rate), but more skill required")

# Tile 274: Wintertodt
add_tile(274, "Obtain Wintertodt items", WINTERTODT_HOURS)

# Tile 275: 1x GWD Drop
add_tile(275, "Obtain 1x GWD Drop", GWD_1X_HOURS)

# Tile 276: 2x Raid Drops
add_tile(276, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 277: 1x DT2 Boss Drop
add_tile(277, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 278: 3x Ecumenical Key
add_tile(278, "Obtain 3x Ecumenical Key", ECUMENICAL_3X_HOURS)

# Tile 279: 1x Awaken DT2 Boss KC (just need to kill one awakened DT2 boss)
# Awakened DT2 bosses take ~5-15 min per kill depending on boss and skill
add_tile(279, "Obtain 1x Awakened DT2 Boss KC", 0.25, "Just 1 kill, any blorva havers?")

# Tile 280: 1x Wildy Boss Ring
p = 1/716; kph = 50
med = ceil(log(0.5) / log(1 - p))
add_tile(280, "Obtain 1x Wildy Boss Ring", med/kph, f"Same as tile 15, 1/716 from singles wildy bosses, {kph} kph, median {med} kc")

# Tile 281: 1x Giant Egg Sack (from Sarachnis, 1/20)
p = 1/20; kph = 67
med = ceil(log(0.5) / log(1 - p))
add_tile(281, "Obtain 1x Giant Egg Sack", med/kph, "1/20 from sarachnis")

# Tile 282: 3x Moons of Peril
add_tile(282, "Obtain 3x Moons of Peril Unique", moons_hours(3))

# Tile 283: 1x Blood/Shadow/Ice/Smoke Quartz (from DT2 area)
# Quartz are all ~1/200, duke is the fastest to kill
p = 1/207; kph = 40
med = ceil(log(0.5) / log(1 - p))
add_tile(283, "Obtain 1x DT2 Quartz", med/kph, f"Assuming 40 duke/hour, median {med} kc")

# Tile 284: Movement
add_movement(284, "Advance to Tile #291", 291)

# Tile 285: 2x Slayer Boss
add_tile(285, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 286: 2x Raid Drops
add_tile(286, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 287: 1x Godsword Shard
p = 3/512; kph = 30
med = ceil(log(0.5) / log(1 - p))
add_tile(287, "Obtain 1x Godsword Shard", med/kph, f"Same as tile 47, 3/512 for any shard, {kph} kph at GWD")

# Tile 288: 3x Crystal Armour Seed
p = 1/50; kph = 6
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(288, "Obtain 3x Crystal Armour Seed", med/kph, f"From CG, median {med} completions")

# Tile 289: 1x Forgotten Lockbox
p = 1/33; kph = 8 # solo
med = ceil(log(0.5) / log(1 - p))
add_tile(289, "Obtain 1x Forgotten Lockbox", med/kph, f"Same as tile 124, 1/33 from solo yama, {kph} kph, median {med} kc, duo rate should be similar")

# Tile 290: 3x Dragon Boots
p = 1/128; kph = 180  # blowpiping nex spiritual mages
med = int(ceil(nbinom.ppf(0.5, 3, p)))
add_tile(290, "Obtain 3x Dragon Boots", med/kph, f"Same as tile 130, 1/128 from Nex Spiritual Mages, {kph} kph, median {med} kc")

# Tile 291: SIT (Lose -1 SKIP)
add_free(291, "Lose -1 SKIP - Roll Again (SIT)")

# Tile 292: 5x Barrows
add_tile(292, "Obtain 5x Barrows Unique", BARROWS_5X_HOURS)

# Tile 293: 1x Colo Drop (any) (Fortis Colosseum)
add_tile(293, "Obtain 1x Colosseum Drop", 3.75, "Slightly faster than echo crystal, but realistically an echo crystal")

# Tile 294: 2x Raid Drops
add_tile(294, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 295: 1x Hueycoatl
add_tile(295, "Obtain 1x Hueycoatl Unique", huey_hours(1))

# Tile 296: 4x Whip or 1x Unsired
add_tile(296, "Obtain 4x Whip or 1x Unsired", WHIP_OR_UNSIRED_4X_HOURS)

# Tile 297: 1x Ballista Component
p = 1/180; kph = 60
med = ceil(log(0.5) / log(1 - p))
add_tile(297, "Obtain 1x Ballista Component", med/kph, f"Same as tile 205, ~1/180 combined from DGs, {kph} kph")

# Tile 298: 1x Holy/Sang/Twisted Kit
p = 1/75; kph = 3
med = ceil(log(0.5) / log(1 - p))
add_tile(298, "Obtain 1x Raid Kit", med/kph, f"Same as tile 273, Assuming 3 CM's/hour, {med} kc on average. HMT would be faster, but more skill required")

# Tile 299: 1x SRA Piece
p = 1/720; kph = 40 # Duke numbers
med = ceil(log(0.5) / log(1 - p))
add_tile(299, "Obtain 1x SRA Piece", med/kph, f"Same as tile 227, Uses duke numbers, 40 kph 1/720")

# Tile 300: 2x Slayer Boss
add_tile(300, "Obtain 2x Slayer Boss Drop", SLAYER_BOSS_2X_HOURS)

# Tile 301: Movement
add_movement(301, "Go back to Tile #290", 290)

# Tile 302: 1x Pet (no chompy/skotizo)
# Chaos ele?
p = 1/300; kph = 48
med = ceil(log(0.5) / log(1 - p))
add_tile(302, "Obtain 1x Pet (no chompy/skotizo)", 11.6, "Chaos elemental, 1/300 48 kph. Low confidence because there could be a better option", confidence="low")

# Tile 303: 1x DT2 Boss Drop
add_tile(303, "Obtain 1x DT2 Boss Drop", DT2_1X_HOURS)

# Tile 304: 2x Raid Drops
add_tile(304, "Obtain 2x Raid Drop", RAID_2X_HOURS)

# Tile 305: 1x Sigil or Holy Elixir (from Corp)
# More restrictive than tile 243 (no spirit shield)
# Holy Elixir 1/171, Sigils combined ~1/585
# Combined: ~1/132
p = 1/132; kph = 7
med = ceil(log(0.5) / log(1 - p))
add_tile(305, "Obtain 1x Sigil/Holy Elixir", med/kph, f"~1/132 from Corp, {kph} kph")

# Tile 306: 1x Med Clue Boots
p = 5/238.6; kph = 10 # 10 clues/hour from eclectics
med = ceil(log(0.5) / log(1 - p))
add_tile(306, "Obtain 1x Med Clue Boots", med/kph, f"same as tile 136, 5/238.6 combined from med caskets, {kph} clues/hr, median {med} caskets")

# Tile 307: 5x Inferno Capes
# Each Inferno run takes 60-90 min for experienced players. Not guaranteed completion.
# Let's say 75 min average with occasional deaths -> ~90 min per cape effective
# 5 * 90 = 450 min = 7.5 hr
add_tile(307, "Obtain 5x Inferno Capes", 7.5, "~90 min per successful run for experienced players")

# Tile 308: 1x Nightmare Unique
p = 1/113; kph = 9 # https://oldschool.runescape.wiki/w/Phosani%27s_Nightmare#Uniques
med = ceil(log(0.5) / log(1 - p))
add_tile(308, "Obtain 1x Nightmare Unique", 11.0, "Same as tile 224, Phosani's ~1/113 combined unique chance, 9 kph")

# Tile 309: 5x Quivers (from... Fortis Colosseum?)
add_tile(309, "Obtain 5x Quivers", 3.0, "40 minute colo's")

# ============================================================
# Now build the Excel file
# ============================================================

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tile Estimates"

# Headers
headers = ["Tile #", "Description", "Category", "Median Hours", "Confidence", "Notes"]
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 60

# Color coding for confidence
conf_fills = {
    "high": PatternFill("solid", fgColor="C6EFCE"),      # green
    "medium": PatternFill("solid", fgColor="FFEB9C"),     # yellow
    "low": PatternFill("solid", fgColor="FFC7CE"),        # red/pink
    "n/a": PatternFill("solid", fgColor="D9E2F3"),        # light blue (movement/free)
}

# Category fills for the category column
cat_fills = {
    "obtain": PatternFill("solid", fgColor="FFFFFF"),
    "movement": PatternFill("solid", fgColor="B4C6E7"),
    "free": PatternFill("solid", fgColor="A9D18E"),
}

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# Sort tiles by number
tiles.sort(key=lambda x: x["tile"])

# Write data
for i, t in enumerate(tiles):
    row = i + 2
    ws.cell(row=row, column=1, value=t["tile"]).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=2, value=t["description"])
    ws.cell(row=row, column=3, value=t["category"].title())
    ws.cell(row=row, column=4, value=t["median_hours"]).number_format = '0.00'
    ws.cell(row=row, column=5, value=t["confidence"].upper())
    ws.cell(row=row, column=6, value=t["notes"])
    
    # Apply confidence color to the whole row
    conf = t["confidence"]
    fill = conf_fills.get(conf, PatternFill())
    cat_fill = cat_fills.get(t["category"], PatternFill())
    
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.font = Font(name="Arial", size=10)
        if col == 5:
            cell.fill = fill
        elif col == 3:
            cell.fill = cat_fill

ws3 = wb.create_sheet("Slayer Bosses")
sum_headers = ["Boss", "Unique Rate", "EHB", "Median KC for Unique", "Hours -> 1 unique", "Hours -> 2 uniques"]
for col, h in enumerate(sum_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="843C0C")
    cell.alignment = header_align
for i, boss in enumerate(SLAYER_BOSSES):
    row = i + 2
    ws3.cell(row=row, column=1, value=boss.name)
    ws3.cell(row=row, column=2, value=boss.unique_rate).number_format = '0.00000'
    ws3.cell(row=row, column=3, value=boss.ehb)
    ws3.cell(row=row, column=4, value=boss.median_kc)
    ws3.cell(row=row, column=5, value=boss.hours_to_unique).number_format = '0.00'
    ws3.cell(row=row, column=6, value=boss.hours_for_two_uniques()).number_format = '0.00'
ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 16
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 16

ws4 = wb.create_sheet("Raids")
sum_headers = ["Raid", "Unique Rate", "EHB", "Median KC for Unique", "Hours -> 1 unique", "Hours -> 2 uniques"]
for col, h in enumerate(sum_headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="843C0C")
    cell.alignment = header_align
for i, raid in enumerate(RAIDS):
    row = i + 2
    ws4.cell(row=row, column=1, value=raid.name)
    ws4.cell(row=row, column=2, value=raid.unique_rate).number_format = '0.00000'
    ws4.cell(row=row, column=3, value=raid.ehb)
    ws4.cell(row=row, column=4, value=raid.median_kc)
    ws4.cell(row=row, column=5, value=raid.hours_to_unique).number_format = '0.00'
    ws4.cell(row=row, column=6, value=raid.hours_for_two_uniques()).number_format = '0.00'
ws4.column_dimensions['A'].width = 32
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 16
ws4.column_dimensions['F'].width = 16


# Add summary sheet
ws2 = wb.create_sheet("Skip Analysis")

# Find only "obtain" tiles, sorted by median hours descending (best skip candidates)
obtain_tiles = [t for t in tiles if t["category"] == "obtain"]
obtain_tiles.sort(key=lambda x: x["median_hours"], reverse=True)

# Headers for summary
sum_headers = ["Rank", "Tile #", "Description", "Median Hours", "Confidence", "Skip Priority"]
for col, h in enumerate(sum_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill("solid", fgColor="843C0C")
    cell.alignment = header_align

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 55
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 15

for i, t in enumerate(obtain_tiles):
    row = i + 2
    ws2.cell(row=row, column=1, value=i+1).alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=2, value=t["tile"]).alignment = Alignment(horizontal="center")
    ws2.cell(row=row, column=3, value=t["description"])
    ws2.cell(row=row, column=4, value=t["median_hours"]).number_format = '0.00'
    ws2.cell(row=row, column=5, value=t["confidence"].upper())
    
    # Skip priority
    if i < 3:
        priority = "TOP 3 SKIP"
        pfill = PatternFill("solid", fgColor="FF0000")
        pfont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    elif i < 10:
        priority = "Strong candidate"
        pfill = PatternFill("solid", fgColor="FFC000")
        pfont = Font(name="Arial", size=10)
    elif i < 20:
        priority = "Consider"
        pfill = PatternFill("solid", fgColor="FFEB9C")
        pfont = Font(name="Arial", size=10)
    else:
        priority = ""
        pfill = PatternFill()
        pfont = Font(name="Arial", size=10)
    
    pcell = ws2.cell(row=row, column=6, value=priority)
    pcell.fill = pfill
    pcell.font = pfont
    
    conf_fill = conf_fills.get(t["confidence"], PatternFill())
    ws2.cell(row=row, column=5).fill = conf_fill
    
    for col in range(1, 7):
        ws2.cell(row=row, column=col).border = thin_border
        if col != 5 and col != 6:
            ws2.cell(row=row, column=col).font = Font(name="Arial", size=10)

# Freeze panes
ws.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'
ws4.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f"A1:F{len(tiles)+1}"
ws2.auto_filter.ref = f"A1:F{len(obtain_tiles)+1}"
ws3.auto_filter.ref = f"A1:F{len(SLAYER_BOSSES)+1}"
ws4.auto_filter.ref = f"A1:F{len(RAIDS)+1}"

output_path = "./snakes_ladders_estimates.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Total tiles: {len(tiles)}")
print(f"Obtain tiles: {len(obtain_tiles)}")
print(f"\nTop 10 skip candidates:")
for i, t in enumerate(obtain_tiles[:10]):
    print(f"  {i+1}. Tile {t['tile']}: {t['description']} - {t['median_hours']:.1f} hrs ({t['confidence']})")
